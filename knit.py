from PIL import Image
from openpyxl import Workbook
#from openpyxl.styles import PatternFill
#from openpyxl.styles.borders import Border, Side
import math
#import pdfplumber
class Colowork_Chart:
    def __init__(self, img, style = "intarsia") -> None:
        self.img = img.convert('RGB')
        self.width = img.width
        self.height = img.height
        self.style = style
        self.pixel_map = self.img_to_map(self.img)

    def img_to_map(self):
        pixel_map = []
        for x in range(self.img.width):
            row = []
            for y in range(self.img.height):
                r, g, b = self.img.getpixel((x, y))
                rgb = '{:02x}{:02x}{:02x}'.format(r, g, b)
                row.append(rgb)
            pixel_map.append(row)
        return pixel_map

    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for x in range(self.width):
            for y in range(self.height):
                colorFill = PatternFill(start_color=self.pixel_map[x][y], end_color=self.pixel_map[x][y],
                                        fill_type='solid')
                ws.cell(y + 1, x + 1).fill = colorFill
                ws.cell(y + 1, x + 1).border = thin_border
            ws.cell(x + 1, y + 2).value = self.height - x

        return wb

    # calculate yarn for this section (intarsia or todo:stranded)
    # needle size, in mm
    # returns a dict with hex color to yarn needed
    def calc_yarn_chart(self, needle_size):
        # first create a dictionary with colors and no. sts
        colors = {}
        for row in self.pixel_map:
            for pixel in row:
                if pixel not in colors:
                    colors[pixel] = 1
                else:
                    colors[pixel] += 1
        # now calculate yarn in meters, estimating one wraparound as the circumference of the needle
        yarn_needed = {}
        perst = needle_size * math.pi
        for color in colors:
            yarn_needed[color] = colors[color] * perst / 1000

        return yarn_needed


class Yarn:
    # weight is fingerling, sport, aran, chunky
    # default width but can change
    # yardage is the length of yarn per ball and also price
    def __init__(self, name, material, yardage, price, weightperball, weight='aran', gauge=None) -> None:
        self.name = name
        self.weight = weight
        self.yardage = yardage
        self.material = material
        self.price = price
        self.weightperball = weightperball
        # a default gauge if not provided
        default_gauges = {'aran': (17, 22), 'bulky': (12, 14), 'dk': (21, 28), 'fingerling': (24, 32)}
        if gauge:
            self.gauge = gauge
        else:
            self.gauge = default_gauges[self.weight]

    def __str__(self) -> str:
        return self.name + " is a(n) " + self.material + self.weight + " yarn.  A " + self.weightperball + " ball costs " + self.price + "."


class Instruction:
    # kind of like a linked list
    # sts is separated by spaces
    # connected if it is connected to another Instruction
    def __init__(self, input_str="", num_repeat=1, sts="", side="RS", name='NA', dec_alt=1) -> None:
        self.input_str = input_str
        self.num_repeat = num_repeat
        self.sts = sts
        self.side = side
        self.name = name
        self.dec_alt = dec_alt
        self.dec_alt_working = dec_alt
        self.connected = None
        self.previous = None
        self.next = None

    def __str__(self) -> str:
        return self.sts

    # do ssk thing later

# create a 2d list from Instructions
# first_inst is the beginning Instruction
def chart(first_inst, total_st = 10):
    sts_map = {'K': '*', 'P': '-', 'SSK': '\\', '2tog': '/', 'PAT': 'PAT'}
    stchart = []
    curr_inst = first_inst

    # Instructions in progress, add the next and connecting Instructions after one is finished
    instr_in_pr = [curr_inst]
    while len(instr_in_pr) != 0:
        #for item in instr_in_pr:
          #  print(item, item.dec_alt, item.num_repeat, end = " ")
        #print()

        curr_row = ['PAT'] * total_st
        for instr in instr_in_pr:
            # for decreasing on alternate rows
            if instr.dec_alt_working != instr.dec_alt:
                instr.dec_alt_working +=1
                continue
            elif instr.dec_alt_working == instr.dec_alt:
                instr.dec_alt_working = 1

            stitches = instr.sts.split(' ')
            # if on the left side
            if stitches[0] == 'PAT':
                instr.num_repeat -=1
                idx = len(curr_row) - 1
                for i in range(len(stitches)):
                    curr_row[idx] = sts_map[stitches[i]]
                    if stitches[i] == 'SSK' or stitches[i] == '2tog':
                        idx-=1
                        curr_row[idx] = sts_map[stitches[i]]
                        total_st -=1
                    idx -=1

            # if on the right side
            if stitches[-1] == 'PAT':
                instr.num_repeat -= 1
                idx = 0
                for i in range(len(stitches)):
                    curr_row[idx] = sts_map[stitches[i]]
                    if stitches[i] == 'SSK' or stitches[i] == '2tog':
                        idx+=1
                        curr_row[idx] = sts_map[stitches[i]]
                        total_st -=1
                    idx += 1

            # append the row to the chart
            stchart.append(curr_row)


            # if the repeats are finished, remove the Instruction
            # and add the next Instruction and any connected Instructions
            if instr.num_repeat == 0:
                instr_in_pr.remove(instr)
                #print(instr, instr.dec_alt)
                if instr.next != None:
                    instr_in_pr.append(instr.next)

                    if instr.next.connected != None and instr.next.connected not in instr_in_pr:
                        instr_in_pr.append(instr.next.connected)

            #curr_row = ['PAT'] * total_st

    return stchart

def chart_to_excel(stchart):
    wb = Workbook()
    ws = wb.active
    first_row_center = len(stchart[0])//2
    print(first_row_center)

    for i in range(len(stchart)):
        start = first_row_center - len(stchart[i])//2
        idx = 0
        for j in range(start, start + len(stchart[i])):
            ws.cell(len(stchart) - i + 1, j+1).value = stchart[i][idx]
            idx+=1

    return wb

def main():
    I1 = Instruction(input_str='3rd row: K1. Sl1. K1. psso. Pat to end of row.', sts='K SSK PAT', name='3rd row')
    I2 = Instruction(input_str='Keeping cont of pat, dec 1 st at raglan edge as before on every following alt row from previous dec 20 times more', num_repeat=20, sts='2tog PAT', dec_alt= 2)
    I1.next = I2
    I3 = Instruction(input_str='then every row 11 times', num_repeat= 11, sts = '2tog PAT', dec_alt = 1 )
    I2.next = I3
    I4 = Instruction(input_str='dec 1 st at V-neck edge on every following 6th row from previous dec 9 times more', num_repeat=9, dec_alt = 6, sts = 'PAT SSK')
    I2.connected = I4
    I3.connected = I4
    I4.connected = I2
    I5 = Instruction(input_str='Dec 1 st at raglan edge, as before, on every row 15 times more.', num_repeat=15, sts = '2tog PAT')
    I4.next = I5

    stchart = chart(I1, 70)
    w = chart_to_excel(stchart)
    w.save('/Users/kingbananacles/Desktop/idk.xlsx')

if __name__ == '__main__':
    main()
